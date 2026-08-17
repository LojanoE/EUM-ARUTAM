# Extrae estudiantes, horarios y tutores de los Excel de DOC/ y genera seed/*.json
# Uso: python scripts/extraer_excel.py
import json
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DOC = ROOT / "DOC"
SEED = ROOT / "seed"

FUENTES = [
    DOC / "1.-ASISTENCIA_EGB_ARUTAM_V0 (2).xlsm",
    DOC / "2.-ASISTENCIA_BACHILLERATO_ARUTAM_V0 (2).xlsm",
]

DIAS = {"LUNES", "MARTES", "MIERCOLES", "MIÉRCOLES", "JUEVES", "VIERNES"}


def limpiar(texto):
    """Quita espacios y corrige caracteres rotos de codificación."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFC", str(texto)).strip()
    # Normaliza comillas tipográficas a comillas rectas (claves más limpias)
    t = t.replace("“", '"').replace("”", '"')
    return " ".join(t.split())


def cargar(ruta):
    return openpyxl.load_workbook(ruta, read_only=True, data_only=True)


def extraer_estudiantes(wb):
    nombre_hoja = next(s for s in wb.sheetnames if s.strip() == "DB_ESTUDIANTES")
    ws = wb[nombre_hoja]
    estudiantes = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # encabezado ID | NOMBRE | GRADO
            continue
        num_id, nombre, grado = (row + (None, None, None))[:3]
        nombre, grado = limpiar(nombre), limpiar(grado)
        if not nombre or not grado:
            continue
        estudiantes.append({"numId": int(num_id), "nombre": nombre, "grado": grado})
    return estudiantes


def extraer_horarios(wb):
    ws = wb["DB_HORARIO"]
    horarios = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # encabezado ID | ASIGNATURA | TIEMPO | DOCENTE | DIA | GRADO
            continue
        _, asignatura, tiempo, docente, dia, grado = (row + (None,) * 6)[:6]
        asignatura, tiempo, docente = limpiar(asignatura), limpiar(tiempo), limpiar(docente)
        dia, grado = limpiar(dia).upper(), limpiar(grado)
        if dia == "MIERCOLES":
            dia = "MIÉRCOLES"
        if not asignatura or dia not in DIAS or not grado:
            continue
        horarios.append({
            "grado": grado,
            "dia": dia,
            "asignatura": asignatura,
            "tiempo": tiempo,
            "docente": docente,
        })
    # orden = posición dentro de cada grado+día (según orden de aparición)
    contadores = {}
    for h in horarios:
        clave = (h["grado"], h["dia"])
        contadores[clave] = contadores.get(clave, 0) + 1
        h["orden"] = contadores[clave]
    return horarios


def extraer_tutores(wb):
    ws = wb["TUTOR_DIAS"]
    tutores = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # encabezado GRADOS | SECCION | TUTOR | ...
            continue
        grado, seccion, tutor = limpiar(row[0]), limpiar(row[1]), limpiar(row[2])
        if grado and tutor:
            tutores[grado] = {"grado": grado, "seccion": seccion or "VESPERTINA", "tutor": tutor}
    return list(tutores.values())


def main():
    SEED.mkdir(exist_ok=True)
    estudiantes, horarios, tutores = [], [], {}
    for ruta in FUENTES:
        print(f"Leyendo {ruta.name} ...")
        wb = cargar(ruta)
        estudiantes += extraer_estudiantes(wb)
        horarios += extraer_horarios(wb)
        for t in extraer_tutores(wb):
            tutores[t["grado"]] = t

    grados_est = sorted({e["grado"] for e in estudiantes})
    grados_hor = sorted({h["grado"] for h in horarios})
    print(f"Estudiantes: {len(estudiantes)} en grados {grados_est}")
    print(f"Horarios:    {len(horarios)} bloques en grados {grados_hor}")
    print(f"Tutores:     {sorted(tutores)}")

    (SEED / "estudiantes.json").write_text(
        json.dumps(estudiantes, ensure_ascii=False, indent=2), encoding="utf-8")
    (SEED / "horarios.json").write_text(
        json.dumps(horarios, ensure_ascii=False, indent=2), encoding="utf-8")
    (SEED / "tutores.json").write_text(
        json.dumps(list(tutores.values()), ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"JSON generados en {SEED}")


if __name__ == "__main__":
    main()
